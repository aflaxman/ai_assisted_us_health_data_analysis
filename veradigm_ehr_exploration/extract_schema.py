"""Turn the delivery's data dictionary spreadsheet into a machine-readable config.

The dictionary is confidential but not limited-use, so it lives outside the repo,
and so does the config this writes. Everything downstream (simulator, inventory
script, notebooks) reads that config through ``schema_config``, so the field
lists and fill rates never get retyped by hand.

Which workbook to read, which sheets it uses and how its sections are labelled
are all properties of the dictionary rather than of this script, so they come
from the ``dictionary_layout`` block of the existing config. The ``roles`` block
is preserved across runs.

    python extract_schema.py [--dictionary PATH] [--out config.local.json]
"""

import argparse
import json
import os
import re

import openpyxl

import schema_config

# Section header rows ("<Entity> Table: demographic information") put prose in
# the field-name column and leave the type column empty; real field rows always
# have a type. Detecting on the empty type is more robust than matching the
# header prose, whose wording is not consistent between sections.


def _clean(value):
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text or None


def parse_field_type(field_type):
    """Split 'varchar (50)' / 'numeric(38,0)' into a base type and size."""
    if not field_type:
        return {"base": None, "raw": field_type}
    match = re.match(r"\s*([A-Za-z_]+)\s*(?:\(([^)]*)\))?", field_type)
    if not match:
        return {"base": None, "raw": field_type}
    base = match.group(1).lower()
    args = match.group(2)
    out = {"base": base, "raw": field_type}
    if args:
        parts = [p.strip() for p in args.split(",")]
        if base in ("varchar", "char"):
            out["length"] = int(parts[0]) if parts[0].isdigit() else None
        elif base in ("numeric", "decimal"):
            out["precision"] = int(parts[0]) if parts[0].isdigit() else None
            out["scale"] = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else None
    return out


def extract_tables(workbook, layout):
    sheet = workbook[layout["schema_sheet"]]
    header_label = layout.get("table_header_label", "Table")
    relabel = layout.get("section_header_relabel", {})
    tables = {}
    order = []
    for row in sheet.iter_rows(values_only=True):
        table, field, ftype, description, status, fill = (list(row) + [None] * 6)[:6]
        table = _clean(table)
        field = _clean(field)
        if not table or not field or table == header_label:
            continue

        # Section header rows carry the table's own description.
        if ftype is None:
            note = field.split(":", 1)[1].strip() if ":" in field else field
            # A few sections are labelled with the wrong name in the dictionary;
            # ``section_header_relabel`` says where those fields really belong.
            key = relabel.get(table, table).replace(" ", "_")
            entry = tables.setdefault(key, {"name": key, "description": None, "fields": []})
            entry["description"] = note
            if key not in order:
                order.append(key)
            continue

        key = table.replace(" ", "_")
        entry = tables.setdefault(key, {"name": key, "description": None, "fields": []})
        if key not in order:
            order.append(key)
        try:
            fill_rate = float(fill)
        except (TypeError, ValueError):
            fill_rate = None
        entry["fields"].append(
            {
                "name": field,
                "type": parse_field_type(_clean(ftype)),
                "description": _clean(description),
                "status": _clean(status),
                "fill_rate": fill_rate,
            }
        )
    return [tables[k] for k in order if tables[k]["fields"]]


def extract_value_sets(workbook, layout):
    header_label = layout.get("value_header_label", "Field Value")
    value_sets = {}
    for key, sheet_name in layout["value_sheets"].items():
        sheet = workbook[sheet_name]
        values = []
        seen_header = False
        for row in sheet.iter_rows(values_only=True):
            value, definition = (list(row) + [None, None])[:2]
            value = _clean(value)
            definition = _clean(definition)
            if not value:
                continue
            if value == header_label:
                seen_header = True
                continue
            if not seen_header:
                continue
            values.append({"value": value, "definition": definition})
        value_sets[key] = values
    return value_sets


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dictionary", default=None,
                        help="workbook to read; defaults to the configured path")
    parser.add_argument("--out", default=schema_config.DEFAULT_PATH)
    parser.add_argument("--layout", default=None,
                        help="config supplying dictionary_layout; defaults to --out")
    args = parser.parse_args()

    source = args.layout or args.out
    try:
        previous = schema_config.read_raw(source)
    except schema_config.ConfigError as exc:
        raise SystemExit(
            f"{exc}\n\n"
            "This script reads that file before writing it: it needs the "
            "'dictionary_layout' block to know which sheets to read, and preserves "
            "the 'roles' block. Copy config.example.json to it and fill in the "
            "layout, or point --layout at a config that already has one."
        ) from None
    layout = previous.get("dictionary_layout")
    if not layout or not layout.get("schema_sheet") or not layout.get("value_sheets"):
        raise SystemExit(
            f"{args.layout or args.out} has no usable 'dictionary_layout' block. "
            "It must name the sheet holding the field list and the sheet for each "
            "documented value set; see config.example.json.")

    dictionary = args.dictionary or layout.get("dictionary_path")
    if not dictionary:
        raise SystemExit("no workbook to read: pass --dictionary or set "
                         "dictionary_layout.dictionary_path in the config")

    workbook = openpyxl.load_workbook(dictionary, read_only=True, data_only=True)
    schema = {
        "source": os.path.basename(dictionary),
        "dictionary_layout": layout,
        # Role assignments and the de-identification thresholds are
        # hand-maintained; a re-extract must not lose them.
        "deidentification": previous.get("deidentification", {}),
        "tables": extract_tables(workbook, layout),
        "value_sets": extract_value_sets(workbook, layout),
        "roles": previous.get("roles", {}),
    }
    with open(args.out, "w") as handle:
        json.dump(schema, handle, indent=2)

    total_fields = sum(len(t["fields"]) for t in schema["tables"])
    print(f"{len(schema['tables'])} tables, {total_fields} fields -> {args.out}")
    for table in schema["tables"]:
        populated = sum(1 for f in table["fields"] if (f["fill_rate"] or 0) > 0)
        print(f"  {table['name']:20s} {len(table['fields']):3d} fields, {populated:3d} populated")

    if not schema["roles"]:
        print("\nno 'roles' block: add one before any downstream script can run")


if __name__ == "__main__":
    main()
